import pathlib
from functools import lru_cache

import openpyxl
from loguru import logger
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from schedule_service.services.vuc_schedule_parser.parser.exceptions import (
    WorksheetCount,
)
from schedule_service.services.vuc_schedule_parser.parser.schemas import (
    Subject,
    WeekSchedule,
    WeekScheduleResult,
    ScheduleResult,
    Platoon,
    WeekDate,
    Day,
    WorkbookSettings,
)


class ScheduleParser:
    def __init__(
        self,
        workbook_path: pathlib.Path,
        workbook_settings: dict[str, WorkbookSettings],
        workbook_settings_path: pathlib.Path = None,
        auto_load=True,
        **kwargs,
    ):
        self.workbook: pathlib.Path | openpyxl.Workbook = workbook_path
        self.workbook_path = workbook_path
        self.workbook_settings_path = workbook_settings_path
        self.workbook_settings = workbook_settings
        if auto_load:
            self.load()

    def __repr__(self):
        return f"ScheduleParser(workbook_path={self.workbook_path.name}, workbook_settings_path={self.workbook_settings_path.name})"

    def __enter__(self):
        self.load()
        return self

    def __exit__(self, *args, **kwargs):
        self.close()
        return args, kwargs

    def load(self):
        self.workbook = openpyxl.load_workbook(self.workbook)
        wb_dir = self.workbook_path.parent.name
        if len(self.workbook.worksheets) > 1:
            raise WorksheetCount(f"Workbook {wb_dir} contain more than 1 sheets")
        logger.info(f"Workbook {wb_dir} has been loaded")
        return self

    def close(self):
        if isinstance(self.workbook, openpyxl.Workbook):
            self.workbook.close()
            return True
        else:
            raise TypeError(self.workbook)

    @lru_cache
    def _workbook_settings(self, sheet_name: str | Worksheet = None):
        if not sheet_name:
            return self.workbook_settings.get(self.workbook.active.title)
        if isinstance(sheet_name, Worksheet):
            sheet_name = sheet_name.title
        return self.workbook_settings.get(sheet_name)

    @staticmethod
    def get_weeks(week: int | tuple[int, int] = (1, 18)):
        weeks = None
        if isinstance(week, tuple):
            weeks = [f"{x} неделя" for x in range(week[0], week[1] + 1)]
        elif isinstance(week, int):
            weeks = [f"{week} неделя"]
        return weeks

    def get_year_range(self, sheet: Worksheet):
        year: list[str] = (
            str(
                sheet.cell(
                    row=self._workbook_settings(sheet).year_range.rows,
                    column=self._workbook_settings(sheet).year_range.cols,
                ).value
            )
            .split(" ")[-3]
            .split("/")
        )
        return int(year[0]), int(year[1])

    @staticmethod
    def find_in_sheet(
        sheet: Worksheet,
        data,
        query: list[str],
    ):
        result: list[Cell] = []
        for d in data:
            for cell in d:
                if cell.value in query:
                    result.append(cell)
        return result

    @staticmethod
    def parse_merged_cell(
        sheet: Worksheet,
        row,
        col,
    ) -> Cell:
        cell = sheet.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    cell = sheet.cell(
                        row=merged_range.min_row, column=merged_range.min_col
                    )
                    if not cell.value:
                        cell = sheet.cell(
                            row=merged_range.max_row, column=merged_range.max_col
                        )
                    break
        return cell

    def find_empty_cell_value(
        self,
        sheet: Worksheet,
        cell: Cell,
        max_col: int,
    ) -> Cell:
        for i in range(self._workbook_settings(sheet).platoon_column_number + 1):
            _cell = sheet.cell(cell.row, cell.column + i)
            if _cell.value or _cell.column > max_col:
                return _cell
        return cell

    def transform_cell(
        self,
        sheet: Worksheet,
        cell: Cell | MergedCell,
        max_col: int,
    ):
        if isinstance(cell, MergedCell):
            return self.parse_merged_cell(sheet, cell.row, cell.column)
        elif not cell.value:
            return self.find_empty_cell_value(sheet, cell, max_col)
        return cell

    def platoons(self, speciality_code: int | None = None) -> list[Platoon]:
        sheet = self.workbook.active
        workbook_settings = self._workbook_settings(sheet)
        unique_platoons = set()
        platoon_ranges = self.find_in_sheet(
            sheet=sheet,
            data=sheet.iter_rows(values_only=False),
            query=["взвод"],
        )

        for platoon_row_cell in platoon_ranges:
            platoon_names = list(
                sheet.iter_rows(
                    min_row=platoon_row_cell.row,
                    max_row=platoon_row_cell.row,
                    min_col=sheet.min_column + 1,
                    max_col=sheet.max_column,
                    values_only=True,
                )
            )
            for platoon_name in platoon_names[0]:
                if platoon_name is not None:
                    if str(platoon_name[0:2]).isdecimal():
                        unique_platoons.add(platoon_name)
        unique_platoons = list(sorted(unique_platoons))
        platoons = []
        for unique_platoon in unique_platoons:
            _platoon_number = Platoon.parse_platoon_number(unique_platoon)
            _specialty_code = Platoon.parse_speciality_code(unique_platoon)
            platoon = Platoon(
                platoon_number=_platoon_number, specialty_code=_specialty_code
            )
            logger.debug(f"platoon '{platoon.model_dump()}'")
            if speciality_code:
                if speciality_code == _specialty_code:
                    platoons.append(platoon)
            else:
                platoons.append(platoon)

        return platoons

    def get_days_week(self, platoon: int = None, **kwargs) -> list[WeekDate]:
        sheet = self.workbook.active
        workbook_settings = self._workbook_settings(sheet)
        result: list[WeekDate] = []

        week_ranges = self.find_in_sheet(
            sheet=sheet,
            data=sheet.iter_rows(values_only=False),
            query=self.get_weeks(),
        )
        for week_range in week_ranges:
            week_number = int(str(week_range.value).split(" ")[0])
            week_date = WeekDate(week=week_number, days=[])
            for day_row_num in range(
                week_range.row + 1,
                week_range.row + workbook_settings.week_range_dim.rows + 1,
                workbook_settings.day_range.rows,
            ):
                day_cell = sheet.cell(row=day_row_num, column=week_range.column)
                if not day_cell.value:
                    continue
                day_platoons_cells = list(
                    sheet.iter_cols(
                        min_row=day_row_num + 1,
                        max_row=day_row_num + 1,
                        min_col=week_range.column,
                        max_col=week_range.column
                        + workbook_settings.day_range.cols
                        - workbook_settings.platoon_column_number,
                        values_only=True,
                    )
                )
                platoons = list(
                    filter(
                        lambda c: c[0] if len(c) > 0 else None,
                        day_platoons_cells,
                    )
                )
                day = Day(day=day_cell.value, platoons=[])
                for i, pl in enumerate(platoons):

                    def _log():
                        logger.debug(
                            f"day '{day.day}' platoon_number '{platoon_number}'"
                        )

                    platoon_number = Platoon.parse_platoon_number(pl[0])
                    if platoon:
                        if platoon == platoon_number:
                            _log()
                            day.platoons.append(platoon_number)
                            break
                    else:
                        _log()
                        day.platoons.append(platoon_number)
                if len(day.platoons) > 0:
                    week_date.days.append(day)
            result.append(week_date)
        return result

    def parse_schedule(self, week: int, platoon: int = None, **kwargs):
        if not week:
            raise ValueError("week required")
        sheet = self.workbook.active
        workbook_settings = self._workbook_settings(sheet)
        result: list[WeekScheduleResult] = []

        week_range = self.find_in_sheet(
            sheet=sheet,
            data=sheet.iter_rows(values_only=False),
            query=self.get_weeks(week),
        )[0]

        def _log(
            date=None,
            platoon=None,
            coordinates=None,
            is_debug=False,
            is_success=False,
            **kwargs,
        ):
            msg = (
                f"workbook dir '{self.workbook_path.parent.name}'"
                f" sheet '{sheet.title}'"
                f" week '{week}'"
                f" date '{date}'"
                f" platoon '{platoon}'"
                f" coordinates {coordinates}; "
            )
            msg += " ".join([f"{k} '{v}'" for k, v in kwargs.items()])
            if is_debug:
                logger.debug(msg)
            elif is_success:
                logger.success(msg)
            else:
                logger.info(msg)

        for day_row_num in range(
            week_range.row + 1,
            week_range.row + workbook_settings.week_range_dim.rows + 1,
            workbook_settings.day_range.rows,
        ):
            _day_ranges = list(
                sheet.iter_cols(
                    min_row=day_row_num,
                    max_row=day_row_num + workbook_settings.day_range.rows - 1,
                    min_col=week_range.column,
                    max_col=week_range.column
                    + workbook_settings.day_range.cols
                    - workbook_settings.platoon_column_number,
                    values_only=False,
                )
            )
            day_platoon_schedules = [
                _day_ranges[idx]
                for idx in range(
                    0,
                    len(_day_ranges),
                    workbook_settings.platoon_column_number,
                )
            ]
            day_platoon_col_max = day_platoon_schedules[-1][0].column
            for pl_start_col, _pl_cells in enumerate(day_platoon_schedules):
                pl_cells = []
                for pl_cell in _pl_cells:
                    pl_cells.append(
                        self.transform_cell(
                            sheet,
                            pl_cell,
                            max_col=day_platoon_col_max,
                        )
                    )
                coordinates = (pl_cells[0].coordinate, pl_cells[-1].coordinate)
                date = pl_cells[0].value
                if not date:
                    _log(date=date, coordinates=coordinates, skip=True, is_debug=True)
                    continue
                platoon_name = pl_cells[1].value
                if platoon_name:
                    platoon_number = Platoon.parse_platoon_number(platoon_name)
                    if platoon:
                        if platoon != platoon_number:
                            _log(
                                date=date,
                                platoon=platoon,
                                coordinates=coordinates,
                                platoon_filter=platoon,
                                skip=True,
                                is_debug=True,
                            )
                            continue

                    specialty_code = Platoon.parse_speciality_code(platoon_name)
                    subject_cell_values = tuple(map(lambda c: c.value, pl_cells[2:]))
                    subjects = []
                    for subject_cvi in range(
                        0,
                        len(subject_cell_values),
                        workbook_settings.subject_rows_number,
                    ):
                        lesson = Subject.parse_auditory(
                            subject_cell_values[subject_cvi + 1]
                        )
                        if subject_cell_values[subject_cvi]:
                            subjects.append(
                                Subject(
                                    name=subject_cell_values[subject_cvi],
                                    auditory=subject_cell_values[subject_cvi + 1],
                                    teacher=subject_cell_values[subject_cvi + 2],
                                    lesson=lesson,
                                )
                            )
                    datetime = WeekSchedule.parse_date_raw(
                        date_day_month=date, year_range=self.get_year_range(sheet)
                    )
                    week_schedule = WeekSchedule(
                        datetime=datetime,
                        date=date,
                        subjects=subjects,
                        coordinates=coordinates,
                    )
                    platoon_obj = Platoon(
                        platoon_number=platoon_number,
                        specialty_code=specialty_code,
                    )
                    week_schedule_result = WeekScheduleResult(
                        platoon=platoon_obj,
                        schedule=week_schedule,
                    )
                    if platoon:
                        if platoon == platoon_number:
                            _log(
                                date=date,
                                platoon=platoon_number,
                                coordinates=coordinates,
                                is_success=True,
                            )
                            result.append(week_schedule_result)
                            return result
                    else:
                        result.append(week_schedule_result)
                else:
                    _log(
                        date=date,
                        platoon=platoon_name,
                        coordinates=coordinates,
                        is_debug=True,
                    )
        return result

    def parse_all_schedule(self, platoon: int = None, **kwargs):
        result: list[ScheduleResult] = []
        weeks = self.get_weeks()
        for week in weeks:
            week_number = int(week.split(" ")[0])
            r = self.parse_schedule(week=week_number, platoon=platoon)
            if r:
                result.append(ScheduleResult(week=week_number, schedule=r))
        return result
